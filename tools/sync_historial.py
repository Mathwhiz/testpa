"""
sync_historial.py
Sincroniza el historial de territorios desde el Excel nuevo a Firestore.

Estrategia: borra historial existente por territorio y re-inserta desde Excel.
Solo toca la subcolección 'historial' — no modifica grupos, publicadores, ni config.

Hojas procesadas:
  Grupo 1, Grupo 2, Grupo 3, ' Grupo 4'
      → 2 columnas por territorio, filas alternadas: conductor (str) / fechas (datetime)
  Congregacion
      → sin conductor, solo fechas (2 cols por territorio)
      → Bloque 1: cabeceras en fila 1  (territorios 11, 26, 27...)
      → Bloque 2: cabeceras en fila 16 (territorios 72, 73, 80...)
      → "Ataliva Roca" en fila 29: se crea el doc de territorio si no existe
  Edificios → ignorada

Uso:
    pip install firebase-admin openpyxl
    python sync_historial.py
"""

import firebase_admin
from firebase_admin import credentials, firestore
from openpyxl import load_workbook
from datetime import datetime
import re

# ─── Config ───────────────────────────────────────────────────────────────────

CONGRE_ID  = "sur"
EXCEL_FILE = "Registro de Asignación de Territorio.xlsx"

HOJAS_GRUPO = {
    "Grupo 1": "1",
    "Grupo 2": "2",
    "Grupo 3": "3",
    " Grupo 4": "4",   # el nombre tiene espacio adelante en el Excel
}
CONGRE_SHEET = "Congregacion"
ATALIVA_ID   = "ataliva_roca"

# ─── Helpers ──────────────────────────────────────────────────────────────────

def fmt_fecha(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if s in ("—", "-", "", "None"):
        return None
    # Formato D/M/YYYY o D/M/YY
    m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{2,4})$", s)
    if m:
        d, mo, y = m.groups()
        if len(y) == 2:
            y = "20" + y
        return f"{y}-{mo.zfill(2)}-{d.zfill(2)}"
    # Formato "D - M - YY" o "D-M-YY" (texto en Excel Grupo 4)
    m = re.match(r"(\d{1,2})\s*-\s*(\d{1,2})\s*-\s*(\d{2,4})$", s)
    if m:
        d, mo, y = m.groups()
        if len(y) == 2:
            y = "20" + y
        return f"{y}-{mo.zfill(2)}-{d.zfill(2)}"
    return None

def is_date_like(val):
    """Detecta si un valor es fecha (datetime o string con formato de fecha)."""
    if isinstance(val, datetime):
        return True
    if isinstance(val, str):
        s = val.strip()
        return bool(
            re.match(r"\d{1,2}/\d{1,2}/\d{2,4}$", s) or
            re.match(r"\d{1,2}\s*-\s*\d{1,2}\s*-\s*\d{2,4}$", s)
        )
    return False

def fmt_nombre(val):
    if val is None:
        return None
    s = str(val).strip()
    if not s or s in ("—", "-"):
        return None
    return s.title()

def terr_num(val):
    if val is None:
        return None
    m = re.search(r"(\d+)", str(val))
    return int(m.group(1)) if m else None

# ─── Parser: hojas de Grupo 1-4 ───────────────────────────────────────────────
# Fila 0 = cabeceras ("Numero Terr X", None, "Numero Terr Y", None, ...)
# Filas 1+: alternadas conductor (str) / fechas (datetime)
# 2 columnas por territorio: col_idx = fechaInicio, col_idx+1 = fechaFin

def parse_grupo(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}

    # Detectar columnas de territorio en fila 0
    terr_cols = []
    for i, val in enumerate(rows[0]):
        if val is not None:
            num = terr_num(val)
            if num:
                terr_cols.append((i, num))

    results = {}
    for col_idx, terr_id in terr_cols:
        entries        = []
        pending_cond   = None

        for row in rows[1:]:
            v0 = row[col_idx]         if col_idx     < len(row) else None
            v1 = row[col_idx + 1]     if col_idx + 1 < len(row) else None

            if v0 is None and v1 is None:
                continue

            if is_date_like(v0):
                # Fila de fechas
                inicio = fmt_fecha(v0)
                fin    = fmt_fecha(v1)
                if inicio:
                    entries.append({
                        "conductor":   fmt_nombre(pending_cond),
                        "fechaInicio": inicio,
                        "fechaFin":    fin,
                    })
                pending_cond = None

            elif v0 is not None:
                # Fila de conductor
                if pending_cond is not None:
                    # Conductor sin fecha asociada (anotación sin cierre)
                    entries.append({
                        "conductor":   fmt_nombre(pending_cond),
                        "fechaInicio": None,
                        "fechaFin":    None,
                    })
                pending_cond = v0

        # Flush: conductor al final sin fechas
        if pending_cond is not None:
            entries.append({
                "conductor":   fmt_nombre(pending_cond),
                "fechaInicio": None,
                "fechaFin":    None,
            })

        if entries:
            results[str(terr_id)] = entries

    return results

# ─── Parser: hoja Congregacion ─────────────────────────────────────────────────
# Sin conductor — solo fechas.
# Bloque 1: cabeceras en fila 1, datos en filas 2-15
# Bloque 2: cabeceras en fila 16, datos en filas 17-28
# Ataliva Roca: etiqueta en fila 29, datos en filas 30-38 (cols 0-1)

def parse_congre(ws):
    rows = list(ws.iter_rows(values_only=True))
    results = {}

    def process_block(header_row_idx, data_start, data_end):
        if header_row_idx >= len(rows):
            return
        header = rows[header_row_idx]
        terr_cols = []
        for j, v in enumerate(header):
            if v is not None and isinstance(v, str):
                num = terr_num(v)
                if num and j % 2 == 0:   # cabeceras en columnas pares
                    terr_cols.append((j, num))

        for i in range(data_start, min(data_end, len(rows))):
            row = rows[i]
            for col_idx, terr_id in terr_cols:
                v0 = row[col_idx]     if col_idx     < len(row) else None
                v1 = row[col_idx + 1] if col_idx + 1 < len(row) else None
                if isinstance(v0, datetime):
                    inicio = fmt_fecha(v0)
                    fin    = fmt_fecha(v1)
                    if inicio:
                        results.setdefault(str(terr_id), []).append({
                            "conductor":   None,
                            "fechaInicio": inicio,
                            "fechaFin":    fin,
                        })

    # Bloque 1 y 2
    process_block(header_row_idx=1,  data_start=2,  data_end=16)
    process_block(header_row_idx=16, data_start=17, data_end=29)

    # Ataliva Roca (cols 0-1, filas 30-38)
    ataliva = []
    for i in range(30, min(39, len(rows))):
        row = rows[i]
        v0  = row[0] if len(row) > 0 else None
        v1  = row[1] if len(row) > 1 else None
        if isinstance(v0, datetime):
            inicio = fmt_fecha(v0)
            fin    = fmt_fecha(v1)
            if inicio:
                ataliva.append({
                    "conductor":   None,
                    "fechaInicio": inicio,
                    "fechaFin":    fin,
                })
    if ataliva:
        results[ATALIVA_ID] = ataliva

    return results

# ─── Firestore: reemplazar historial ─────────────────────────────────────────

def replace_historial(terr_ref, entries, label=""):
    existing   = list(terr_ref.collection("historial").stream())
    deleted    = len(existing)
    for doc in existing:
        doc.reference.delete()

    # Insertar en lotes de 400
    batch     = db.batch()
    batch_n   = 0
    total_new = 0
    for entry in entries:
        ref = terr_ref.collection("historial").document()
        batch.set(ref, entry)
        batch_n   += 1
        total_new += 1
        if batch_n == 400:
            batch.commit()
            batch   = db.batch()
            batch_n = 0
    if batch_n:
        batch.commit()

    print(f"   {label}: {deleted} borrados → {total_new} nuevos")
    return total_new

# ─── Main ─────────────────────────────────────────────────────────────────────

import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

print("==> Conectando a Firebase...")
cred = credentials.Certificate("serviceAccountKey.json")
firebase_admin.initialize_app(cred)
db = firestore.client()
congre_ref = db.collection("congregaciones").document(CONGRE_ID)

wb = load_workbook(EXCEL_FILE, read_only=True)

total_terr    = 0
total_entries = 0

# ── Grupos 1-4 ────────────────────────────────────────────────────────────────
print("\n==> Grupos...")
for hoja, grupo_id in HOJAS_GRUPO.items():
    ws   = wb[hoja]
    data = parse_grupo(ws)
    print(f"  Hoja '{hoja.strip()}' (Grupo {grupo_id}): {len(data)} territorios")
    for terr_id_str, entries in data.items():
        terr_ref = congre_ref.collection("territorios").document(terr_id_str)
        n = replace_historial(terr_ref, entries, f"T{terr_id_str}")
        total_terr    += 1
        total_entries += n

# ── Congregacion ──────────────────────────────────────────────────────────────
print("\n==> Congregacion...")
ws_c   = wb[CONGRE_SHEET]
data_c = parse_congre(ws_c)
print(f"  {len(data_c)} territorios encontrados")

for terr_id_str, entries in data_c.items():
    if terr_id_str == ATALIVA_ID:
        # Crear doc de territorio si no existe (sin polígono — no aparece en mapa)
        terr_ref = congre_ref.collection("territorios").document(ATALIVA_ID)
        if not terr_ref.get().exists:
            terr_ref.set({
                "id":       ATALIVA_ID,
                "nombre":   "Ataliva Roca",
                "grupoId":  "C",
                "tipo":     "normal",
                "punto":    None,
                "poligonos": [],
            })
            print(f"  [+] Creado territorio 'Ataliva Roca' en Firestore")
    else:
        terr_ref = congre_ref.collection("territorios").document(terr_id_str)

    n = replace_historial(terr_ref, entries, f"T{terr_id_str}")
    total_terr    += 1
    total_entries += n

# ── Resumen ───────────────────────────────────────────────────────────────────
print(f"\n{'='*50}")
print(f"SYNC COMPLETO")
print(f"   Territorios actualizados : {total_terr}")
print(f"   Entradas de historial    : {total_entries}")
print(f"{'='*50}")
