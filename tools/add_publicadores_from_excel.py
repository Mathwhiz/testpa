"""
add_publicadores_from_excel.py
Extrae todos los nombres individuales del Excel de VM y los agrega a Firestore
como publicadores. Los que ya existen (por normalización) se saltan.

Además usa la hoja "Anc y SM" para etiquetar con VM_PRESIDENTE y VM_ORACION
a los Ancianos, y con otros roles a los Siervos Ministeriales.

Uso:
    python add_publicadores_from_excel.py [--dry-run]
"""

import re
import sys
import io
import unicodedata
from collections import defaultdict

import firebase_admin
from firebase_admin import credentials, firestore
from openpyxl import load_workbook

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

DRY_RUN = "--dry-run" in sys.argv

# ─── Correcciones manuales ─────────────────────────────────────────────────────
# norm(nombre_incorrecto) → nombre_canonico
# Incluir solo los casos que NO se resuelven por normalización de acentos.
CORRECCIONES = {
    # Typos de apellido
    "armado nunez":            "Armando Nuñez",
    "blanca diaz":             "Bianca Diaz",
    "dora madina":             "Dora Medina",
    "elizabeht camarata":      "Elizabeth Camarata",
    "elizabeht diaz":          "Elizabeth Diaz",
    "elizabeth camaratta":     "Elizabeth Camarata",
    "elsa reynoso":            "Elsa Reinoso",
    "feliz villatoro":         "Félix Villatoro",
    "frenando oberts":         "Fernando Oberts",
    "gladis bazan":            "Gladys Bazan",
    "gisele hernandez":        "Gisselle Hernandez",
    "hortencia payes":         "Hortensia Payes",
    "jonathan zurita":         "Jonatan Zurita",
    "jose luis lasierra":      "José Luis Lasierra",
    "jose luis lasierrra":     "José Luis Lasierra",
    "juliana guisetti":        "Juliana Guizzetti",
    "juliana guissetti":       "Juliana Guizzetti",
    "juliana guizeti":         "Juliana Guizzetti",
    "juliana guizetti":        "Juliana Guizzetti",
    "malanie scalese":         "Melanie Scalese",
    "melanie scalece":         "Melanie Scalese",
    "maria herhenreder":       "Maria Hergenreder",
    "mauro tobres":            "Mauro Tobares",
    "miryan d'adam":           "Miryam D'adam",
    "myriam d'adam":           "Miryam D'adam",
    "orar flores":             "Omar Flores",
    "rodrigo bustos":          "Rodrigo Busto",
    "rut carra":               "Ruth Carra",
    "stafano camarata":        "Stefano Camarata",
    "stefano camarada":        "Stefano Camarata",
    "susana ferrrer":          "Susana Ferrer",
    "walther":                 "Walther Gil",
    "analia bustos":           "Analia Busto",
    "isabela camarata":        "Isabella Camarata",
    "ariel oberts":            "Ariel Oberst",
    "alejandra oberts":        "Alejandra Oberst",
    "pamela buenos":           "Pamela Bueno",
    "benjamin bustos":         "Benjamin Busto",
    "catalina basto":          "Catalina Bastos",
    "ana maria llanos":        "Ana Llanos",   # mismo apellido, puede ser misma o distinta persona — agrupar
    "loisa viviana":           "Viviana Loisa",
    "enzo acota":              "Enzo Acosta",
    "luis zorrillla":          "Luis Zorrilla",
    "nelida rodriguezl":       "Nelida Rodriguez",
    # Nombres invertidos
    "zurita jonatan":          "Jonatan Zurita",
    "espinal emmanuel":        "Emmanuel Espinal",
    "fernandez andrea":        "Andrea Fernández",
    "sanchez patricia":        "Patricia Sanchez",
    # Anotaciones en el nombre
    "benjamin oberts-":        "Benjamín Oberts",
    "benjamin carrizo (reemplazo)": "Benjamín Carrizo",
}

# Entradas a excluir totalmente (no son personas)
EXCLUIR = {
    "sala principa",    # artefacto de parsing
    "walther",          # nombre incompleto (ya mapeado arriba, pero por si acaso)
}

# ─── Config ───────────────────────────────────────────────────────────────────

CONGRE_ID  = "sur"
EXCEL_FILE = "Copia de Reunión Vida y Ministerio Cristiano.xlsx"
HOJAS_IGNORAR   = {"Anc y SM"}
SUFIJO_IGNORAR  = "SR"

# ─── Helpers ──────────────────────────────────────────────────────────────────

def normalize(s):
    if not s:
        return ""
    s = str(s).strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFD", s)
                if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", s)


def fmt_nombre(val):
    if val is None:
        return None
    s = str(val).strip()
    return s if s and s not in ("—", "-") else None


def parece_nombre(s):
    """Heurística: ¿parece un nombre de persona? (al menos 3 chars y sin solo números)"""
    if not s:
        return False
    s = s.strip()
    return len(s) >= 3 and not s.isdigit()


def split_todos(val):
    """
    Divide un valor de celda en nombres individuales.
    Maneja: "Name1 - Name2", "Name1 / Name2", "Name1/Name2", "Name1-Name2" (sin espacios).
    Devuelve lista de nombres individuales.
    """
    if not val:
        return []
    s = str(val).strip()
    if not s or s in ("—", "-"):
        return []

    # Separadores con espacios primero
    for sep in [" - ", " / "]:
        if sep in s:
            partes = [p.strip() for p in s.split(sep)]
            return [p for p in partes if parece_nombre(p)]

    # Slash sin espacio
    if "/" in s:
        partes = [p.strip() for p in s.split("/")]
        if all(parece_nombre(p) for p in partes):
            return [p for p in partes if parece_nombre(p)]

    # Dash sin espacios: sólo separar si ambas partes parecen nombres con espacio interno
    if "-" in s:
        partes = [p.strip() for p in s.split("-", 1)]
        if (len(partes) == 2 and parece_nombre(partes[0]) and parece_nombre(partes[1])
                and (" " in partes[0] or " " in partes[1])):
            return [p for p in partes if parece_nombre(p)]

    return [s] if parece_nombre(s) else []


# ─── Extraer nombres de las hojas de datos ────────────────────────────────────

def extraer_nombres_hoja(ws):
    nombres = []
    for row in ws.iter_rows(values_only=True):
        for col_idx in [1, 2]:  # Col B y Col C
            val = row[col_idx] if col_idx < len(row) else None
            if not val:
                continue
            s = str(val).strip()
            # Ignorar sub-headers del Excel
            if s.lower() in ("sala principal", "sala auxiliar"):
                continue
            nombres.extend(split_todos(s))
    return nombres


# ─── Leer hoja Anc y SM ───────────────────────────────────────────────────────

def leer_anc_sm(ws):
    """Devuelve {nombre_normalizado: [roles]}"""
    roles_map = {}
    rows = list(ws.iter_rows(values_only=True))
    for row in rows[1:]:  # skip header
        col_a = fmt_nombre(row[0] if len(row) > 0 else None)
        col_b = fmt_nombre(row[1] if len(row) > 1 else None)
        if col_a:
            roles_map[normalize(col_a)] = ["VM_PRESIDENTE", "VM_ORACION", "VM_TESOROS"]
        if col_b:
            roles_map[normalize(col_b)] = ["VM_MINISTERIO", "VM_VIDA_CRISTIANA"]
    return roles_map


# ─── Main ─────────────────────────────────────────────────────────────────────

print("==> Conectando a Firebase...")
cred = credentials.Certificate("serviceAccountKey.json")
firebase_admin.initialize_app(cred)
db = firestore.client()
congre_ref = db.collection("congregaciones").document(CONGRE_ID)

print(f"\n==> Cargando publicadores existentes de '{CONGRE_ID}'...")
pubs_snap = congre_ref.collection("publicadores").stream()
existentes_norm  = {}  # normalize(nombre) → pubId
existentes_raw   = {}  # normalize(nombre) → nombre raw
for pub in pubs_snap:
    data = pub.to_dict()
    nombre = data.get("nombre", "")
    if nombre:
        k = normalize(nombre)
        existentes_norm[k] = pub.id
        existentes_raw[k]  = nombre
print(f"    {len(existentes_norm)} publicadores existentes.")

print(f"\n==> Abriendo Excel: {EXCEL_FILE}")
wb = load_workbook(EXCEL_FILE, read_only=True, data_only=True)

# Roles de Anc y SM
roles_ancsm = {}
if "Anc y SM" in wb.sheetnames:
    roles_ancsm = leer_anc_sm(wb["Anc y SM"])
    print(f"    {len(roles_ancsm)} personas con roles desde 'Anc y SM'.")

# Recopilar todos los nombres crudos de todas las hojas
# norm_key → lista de apariciones (raw strings)
nombre_apariciones = defaultdict(list)

for nombre_hoja in wb.sheetnames:
    if nombre_hoja in HOJAS_IGNORAR or nombre_hoja.strip().endswith(SUFIJO_IGNORAR):
        continue
    ws = wb[nombre_hoja]
    nombres = extraer_nombres_hoja(ws)
    for n in nombres:
        if parece_nombre(n):
            nombre_apariciones[normalize(n)].append(n)

print(f"\n==> {len(nombre_apariciones)} nombres únicos (normalizados) encontrados en el Excel.")

# Para cada nombre normalizado: elegir la versión "canónica" (más frecuente)
nuevos = []   # (nombre_canonico, norm_key)
ya_hay = []

# Mapa de norm_key → nombre_canonico después de aplicar correcciones
nombre_canonico_map = {}   # norm_key → nombre_canonico final
for norm_key, apariciones in nombre_apariciones.items():
    if norm_key in EXCLUIR:
        continue
    if norm_key in CORRECCIONES:
        nombre_canonico_map[norm_key] = CORRECCIONES[norm_key]
    else:
        # Elegir la versión más común (o la más larga si hay empate)
        conteo = defaultdict(int)
        for a in apariciones:
            conteo[a] += 1
        nombre_canonico_map[norm_key] = max(conteo, key=lambda x: (conteo[x], len(x)))

# Agrupar por nombre_canonico final (las correcciones pueden unificar varios norm_keys)
canonico_a_norm = defaultdict(list)
for norm_key, canonico in nombre_canonico_map.items():
    canonico_a_norm[canonico].append(norm_key)

# norm del canonico final → conjunto de norm_keys que lo representan
# Usamos el norm del canonico como clave de Firestore
for canonico, norm_keys in sorted(canonico_a_norm.items()):
    norm_canonico = normalize(canonico)
    if norm_canonico in existentes_norm:
        ya_hay.append((existentes_raw[norm_canonico], norm_canonico))
        continue
    nuevos.append((canonico, norm_canonico))

print(f"    Ya en Firestore : {len(ya_hay)}")
print(f"    Por agregar     : {len(nuevos)}")

# Mostrar los nuevos
print(f"\n{'─'*55}")
print(f"NUEVOS PUBLICADORES ({len(nuevos)}):")
for nombre, _ in sorted(nuevos):
    print(f"  + {nombre}")

# Subir a Firestore
if not DRY_RUN and nuevos:
    print(f"\n==> Subiendo {len(nuevos)} nuevos publicadores...")
    batch   = db.batch()
    batch_n = 0
    subidos = 0

    for nombre_canonico, norm_key in nuevos:
        roles = roles_ancsm.get(norm_key, [])
        ref = congre_ref.collection("publicadores").document()
        batch.set(ref, {
            "nombre": nombre_canonico,
            "roles":  roles,
            "activo": True,
        })
        batch_n += 1
        subidos += 1
        if batch_n == 400:
            batch.commit()
            batch   = db.batch()
            batch_n = 0

    if batch_n:
        batch.commit()

    print(f"    {subidos} publicadores subidos.")
elif DRY_RUN:
    print("\n[DRY-RUN] No se subió nada a Firestore.")

# ─── Resumen de variantes / posibles duplicados ───────────────────────────────
print(f"\n{'─'*55}")
print("POSIBLES VARIANTES (mismo nombre normalizado con distintas grafías):")
printed = 0
for norm_key, apariciones in sorted(nombre_apariciones.items()):
    unique_raw = list(dict.fromkeys(apariciones))  # preserva orden, deduplica
    if len(unique_raw) > 1:
        print(f"  [{norm_key}]")
        conteo = defaultdict(int)
        for a in apariciones:
            conteo[a] += 1
        for v in sorted(unique_raw, key=lambda x: -conteo[x]):
            print(f"    '{v}' × {conteo[v]}")
        printed += 1
if printed == 0:
    print("  (ninguna variante detectada)")

print(f"\nListo. Ahora podés correr import_vm_historial.py para re-importar con los nuevos pubIds.")
