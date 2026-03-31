"""
import_vm_historial.py
Importa el historial de reuniones VM desde el Excel a Firestore.

Hojas procesadas: todas las que tienen nombre tipo "Mes YY" o "Mes 25/26" etc.
Hojas ignoradas:  las que terminan en "SR", "Anc y SM".

Layout por hoja:
  Fila 1 (idx 0): título de la hoja (ignorado)
  Fila 2 (idx 1): Col A = "Sala Auxiliar: Nombre" → conductorAuxMes
  Bloques verticales con una fila de cabecera:
    Col A = "Semana del DD al DD de Mes YYYY"
    Col A = etiqueta del rol/parte
    Col B = persona asignada (Sala Principal)
    Col C = persona asignada (Sala Auxiliar, cuando aplica)

Schema Firestore destino:
  congregaciones/{congreId}/vidaministerio/{YYYY-MM-DD} →
    fecha, cancionApertura, cancionIntermedia, cancionCierre,
    presidente, oracionApertura, oracionCierre,
    conductorAuxMes,
    tesoros: { discurso, joyas, lecturaBiblica },
    ministerio: [],
    vidaCristiana: [],
    estudioBiblico: {}

Uso:
    pip install firebase-admin openpyxl
    python import_vm_historial.py
"""

import re
import sys
import io
import unicodedata
from datetime import datetime, date, timedelta

import firebase_admin
from firebase_admin import credentials, firestore
from openpyxl import load_workbook

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

# ─── Configuración ────────────────────────────────────────────────────────────

CONGRE_ID  = "sur"
EXCEL_FILE = "Copia de Reunión Vida y Ministerio Cristiano.xlsx"

# Hojas a ignorar
HOJAS_IGNORAR = {"Anc y SM"}
SUFIJO_IGNORAR = "SR"  # cualquier hoja que termine en "SR"

MESES_ES = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4,
    "mayo": 5, "junio": 6, "julio": 7, "agosto": 8,
    "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12,
}

# ─── Helpers ──────────────────────────────────────────────────────────────────

def normalize(s):
    """Normaliza nombre para comparación: minúsculas, sin tildes, sin espacios dobles."""
    if not s:
        return ""
    s = str(s).strip().lower()
    s = "".join(
        c for c in unicodedata.normalize("NFD", s)
        if unicodedata.category(c) != "Mn"
    )
    return re.sub(r"\s+", " ", s)


def fmt_nombre(val):
    if val is None:
        return None
    s = str(val).strip()
    if not s or s in ("—", "-"):
        return None
    return s


def celda(row, col):
    """Lee celda de forma segura (row es tuple de valores)."""
    return row[col] if col < len(row) else None


def parse_cancion(text):
    """Extrae número de canción de strings como 'Canción 52' o 'Canción 52 y oración'."""
    m = re.search(r"Canci[oó]n\s+(\d+)", str(text), re.IGNORECASE)
    return int(m.group(1)) if m else None


def parse_duracion(text):
    """Extrae duración en minutos de strings como '10 mins.' o '(4 min)'."""
    m = re.search(r"(\d+)\s*min", str(text), re.IGNORECASE)
    return int(m.group(1)) if m else None


def parse_semana_str(text):
    """
    Parsea strings de semana en dos formatos:
      "Semana del 10 al 16 de Agosto de 2025"          → mismo mes
      "Semana del 26 de Enero al 01 de Febrero 2026"   → cruza mes
    Devuelve la fecha del lunes de esa semana (YYYY-MM-DD).
    """
    if not text:
        return None
    s = str(text).strip()
    if not re.search(r"Semana\s+del", s, re.IGNORECASE):
        return None

    # Día de inicio: primer número después de "Semana del"
    m_dia = re.search(r"Semana\s+del\s+(\d{1,2})", s, re.IGNORECASE)
    if not m_dia:
        return None
    dia = int(m_dia.group(1))

    # Año: último número de 4 dígitos en el string
    anios = re.findall(r"\b(\d{4})\b", s)
    if not anios:
        return None
    anio = int(anios[-1])

    # Mes: primer nombre de mes que aparece en el string
    s_norm = normalize(s)
    primer_mes = None
    primer_pos = len(s_norm)
    for mes_nombre, mes_num in MESES_ES.items():
        m_pos = s_norm.find(mes_nombre)
        if m_pos != -1 and m_pos < primer_pos:
            primer_pos = m_pos
            primer_mes = mes_num
    if not primer_mes:
        return None

    try:
        d = date(anio, primer_mes, dia)
    except ValueError:
        return None
    # Ajustar al lunes
    if d.weekday() != 0:
        d = d - timedelta(days=d.weekday())
    return d.strftime("%Y-%m-%d")


def parece_nombre(s):
    """Heurística: ¿parece un nombre de persona?"""
    if not s:
        return False
    s = s.strip()
    return len(s) >= 3 and not s.isdigit()


def split_par(val):
    """
    Divide un par "Nombre1 - Nombre2", "Nombre1/Nombre2" o "Nombre1-Nombre2" en [nombre1, nombre2].
    Devuelve [nombre1, None] si no hay separador claro.
    """
    if not val:
        return None, None
    s = str(val).strip()
    if not s or s in ("—", "-"):
        return None, None

    # Separadores con espacios primero
    for sep in [" - ", " / "]:
        if sep in s:
            partes = [p.strip() for p in s.split(sep, 1)]
            return fmt_nombre(partes[0]), fmt_nombre(partes[1]) if len(partes) > 1 else None

    # Slash sin espacio
    if "/" in s:
        partes = [p.strip() for p in s.split("/", 1)]
        if all(parece_nombre(p) for p in partes):
            return fmt_nombre(partes[0]), fmt_nombre(partes[1])

    # Dash sin espacios: separar solo si ambas partes parecen nombres completos
    if "-" in s:
        partes = [p.strip() for p in s.split("-", 1)]
        if (len(partes) == 2 and parece_nombre(partes[0]) and parece_nombre(partes[1])
                and (" " in partes[0] or " " in partes[1])):
            return fmt_nombre(partes[0]), fmt_nombre(partes[1])

    return fmt_nombre(s), None


# ─── Matching de publicadores ─────────────────────────────────────────────────

nombre_to_id: dict = {}
unmatched_log: list = []

# Correcciones: norm(forma incorrecta) → norm(forma canónica en Firestore)
# Sincronizado con add_publicadores_from_excel.py
CORRECCIONES_NORM = {
    "armado nunez":                  "armando nunez",
    "blanca diaz":                   "bianca diaz",
    "dora madina":                   "dora medina",
    "elizabeht camarata":            "elizabeth camarata",
    "elizabeht diaz":                "elizabeth diaz",
    "elizabeth camaratta":           "elizabeth camarata",
    "elsa reynoso":                  "elsa reinoso",
    "feliz villatoro":               "felix villatoro",
    "frenando oberts":               "fernando oberts",
    "gladis bazan":                  "gladys bazan",
    "gisele hernandez":              "gisselle hernandez",
    "hortencia payes":               "hortensia payes",
    "jonathan zurita":               "jonatan zurita",
    "jose luis lasierra":            "jose luis lasierra",
    "jose luis lasierrra":           "jose luis lasierra",
    "juliana guisetti":              "juliana guizzetti",
    "juliana guissetti":             "juliana guizzetti",
    "juliana guizeti":               "juliana guizzetti",
    "juliana guizetti":              "juliana guizzetti",
    "malanie scalese":               "melanie scalese",
    "melanie scalece":               "melanie scalese",
    "maria herhenreder":             "maria hergenreder",
    "mauro tobres":                  "mauro tobares",
    "miryan d'adam":                 "miryam d'adam",
    "myriam d'adam":                 "miryam d'adam",
    "orar flores":                   "omar flores",
    "rodrigo bustos":                "rodrigo busto",
    "rut carra":                     "ruth carra",
    "stafano camarata":              "stefano camarata",
    "stefano camarada":              "stefano camarata",
    "susana ferrrer":                "susana ferrer",
    "walther":                       "walther gil",
    "analia bustos":                 "analia busto",
    "isabela camarata":              "isabella camarata",
    "ariel oberts":                  "ariel oberst",
    "alejandra oberts":              "alejandra oberst",
    "pamela buenos":                 "pamela bueno",
    "benjamin bustos":               "benjamin busto",
    "catalina basto":                "catalina bastos",
    "ana maria llanos":              "ana llanos",
    "loisa viviana":                 "viviana loisa",
    "enzo acota":                    "enzo acosta",
    "luis zorrillla":                "luis zorrilla",
    "nelida rodriguezl":             "nelida rodriguez",
    "zurita jonatan":                "jonatan zurita",
    "espinal emmanuel":              "emmanuel espinal",
    "fernandez andrea":              "andrea fernandez",
    "sanchez patricia":              "patricia sanchez",
    "benjamin oberts-":              "benjamin oberts",
    "benjamin carrizo (reemplazo)":  "benjamin carrizo",
    # Acento extra
    "armando nunez":                 "armando nunez",
}


def match_pub(nombre_raw):
    """Intenta encontrar el pubId de un publicador por nombre. Devuelve None si no lo encuentra."""
    if not nombre_raw:
        return None
    n = normalize(nombre_raw)
    if not n:
        return None

    # Aplicar correcciones de typos conocidos
    n = CORRECCIONES_NORM.get(n, n)

    # Coincidencia exacta
    if n in nombre_to_id:
        return nombre_to_id[n]

    # Coincidencia parcial: todas las palabras del nombre están en el nombre del pub
    partes = n.split()
    for key, pid in nombre_to_id.items():
        kpartes = key.split()
        if all(p in kpartes for p in partes):
            return pid

    # Fallback: si parece un par sin splitear, intentar con la primera mitad
    p1, p2 = split_par(nombre_raw)
    if p2 is not None:
        # Era un par — no loguear el par completo, sí loguear los individuales sin match
        match_pub(p1)
        match_pub(p2)
        return match_pub(p1)  # retorna el pubId del primero (estudiante)

    # No encontrado
    if nombre_raw not in unmatched_log:
        unmatched_log.append(nombre_raw)
    return None


# ─── Parser de bloque semanal ─────────────────────────────────────────────────

TESOROS_MARKER    = re.compile(r"Tesoros\s+de\s+la", re.IGNORECASE)
MINISTERIO_MARKER = re.compile(r"Seamos\s+Mejores\s+Maestros", re.IGNORECASE)
VIDA_MARKER       = re.compile(r"Nuestra\s+Vida\s+Cristiana", re.IGNORECASE)
ORACION_RE        = re.compile(r"^Oraci[oó]n", re.IGNORECASE)
PRESIDENTE_RE     = re.compile(r"Palabras\s+de\s+Introducci", re.IGNORECASE)
CONCLUSION_RE     = re.compile(r"conclusi[oó]n", re.IGNORECASE)
CANCION_RE        = re.compile(r"Canci[oó]n\s+(\d+)", re.IGNORECASE)
NUMBERED_RE       = re.compile(r"^(\d+)\.\s+(?:(\d+)\s*min[s.]?\s+)?(.+)?", re.IGNORECASE)
AUX_MARKER_RE     = re.compile(r"Sala\s+Principal", re.IGNORECASE)


def parse_block(rows):
    """
    Procesa las filas de un bloque semanal y devuelve el dict con el programa.
    rows: lista de tuplas (col_A, col_B, col_C, ...)
    """
    result = {
        "cancionApertura":   None,
        "cancionIntermedia": None,
        "cancionCierre":     None,
        "presidente":        None,
        "oracionApertura":   None,
        "oracionCierre":     None,
        "tesoros": {
            "discurso":       {"titulo": "", "duracion": 10, "pubId": None},
            "joyas":          {"titulo": "", "duracion": 10, "pubId": None},
            "lecturaBiblica": {"titulo": "", "duracion": 4,  "pubId": None, "ayudante": None},
        },
        "ministerio":    [],
        "vidaCristiana": [],
        "estudioBiblico": {"titulo": "", "duracion": 30, "conductor": None, "lector": None},
    }

    INIT, TESOROS, MINISTERIO, VIDA = range(4)
    state          = INIT
    oracion_count  = 0
    tesoros_idx    = 0   # 0=discurso, 1=joyas, 2=lectura
    min_count      = 0
    vida_count     = 0

    for row in rows:
        label_raw = celda(row, 0)
        val_b     = celda(row, 1)
        val_c     = celda(row, 2)

        if label_raw is None and val_b is None:
            continue

        label = str(label_raw).strip() if label_raw is not None else ""

        # ── Marcadores de sección ────────────────────────────────
        if TESOROS_MARKER.search(label):
            state = TESOROS
            continue
        if MINISTERIO_MARKER.search(label):
            state = MINISTERIO
            continue
        if VIDA_MARKER.search(label):
            state = VIDA
            continue

        # ── Sala auxiliar sub-header (None, "Sala Principal", "Sala Auxiliar") ─
        if val_b and AUX_MARKER_RE.search(str(val_b)):
            continue

        # ── Canciones ────────────────────────────────────────────
        m_can = CANCION_RE.search(label)
        if m_can and not NUMBERED_RE.match(label):
            num = int(m_can.group(1))
            if result["cancionApertura"] is None:
                result["cancionApertura"] = num
            elif result["cancionIntermedia"] is None:
                result["cancionIntermedia"] = num
            else:
                result["cancionCierre"] = num
            continue

        # ── Oración ──────────────────────────────────────────────
        if ORACION_RE.match(label) and (val_b or label_raw):
            nombre = match_pub(fmt_nombre(val_b)) if val_b else None
            if oracion_count == 0:
                result["oracionApertura"] = nombre
            else:
                result["oracionCierre"] = nombre
            oracion_count += 1
            continue

        # ── Presidente / palabras de introducción ─────────────────
        if PRESIDENTE_RE.search(label) and val_b:
            result["presidente"] = match_pub(fmt_nombre(val_b))
            continue

        # ── Palabras de conclusión ────────────────────────────────
        if CONCLUSION_RE.search(label) and val_b and result["presidente"] is None:
            result["presidente"] = match_pub(fmt_nombre(val_b))
            continue

        # ── Partes numeradas ──────────────────────────────────────
        m_num = NUMBERED_RE.match(label)
        if m_num:
            num    = int(m_num.group(1))
            dur    = int(m_num.group(2)) if m_num.group(2) else None
            titulo = m_num.group(3).strip() if m_num.group(3) else ""
            # Limpiar el número de párrafo si quedó en el título
            titulo = re.sub(r"^\d+\.\s*", "", titulo).strip()
            pub_b  = match_pub(fmt_nombre(val_b)) if val_b else None

            if state == TESOROS:
                if tesoros_idx == 0:
                    result["tesoros"]["discurso"] = {
                        "titulo": titulo, "duracion": dur or 10, "pubId": pub_b
                    }
                elif tesoros_idx == 1:
                    result["tesoros"]["joyas"] = {
                        "titulo": titulo, "duracion": dur or 10, "pubId": pub_b
                    }
                elif tesoros_idx == 2:
                    # Lectura: Col C = sala auxiliar
                    pub_c = match_pub(fmt_nombre(val_c)) if val_c else None
                    result["tesoros"]["lecturaBiblica"] = {
                        "titulo": titulo, "duracion": dur or 4,
                        "pubId": pub_b, "ayudante": pub_c,
                    }
                tesoros_idx += 1

            elif state == MINISTERIO:
                # Col B tiene "Estudiante - Asistente"
                est_b, asis_b = split_par(val_b)
                # Col C tiene el par de la sala auxiliar
                est_c, asis_c = split_par(val_c) if val_c else (None, None)
                parte = {
                    "titulo":   titulo,
                    "tipo":     "demostracion",
                    "duracion": dur,
                    "pubId":    match_pub(est_b),
                    "ayudante": match_pub(asis_b),
                }
                if est_c:
                    parte["salaAux"] = {
                        "pubId":    match_pub(est_c),
                        "ayudante": match_pub(asis_c),
                    }
                result["ministerio"].append(parte)
                min_count += 1

            elif state == VIDA:
                titulo_lower = titulo.lower()
                es_estudio = (
                    "estudio" in titulo_lower or
                    num >= 9 or
                    vida_count >= 2  # heurística: el estudio es la última parte de Vida
                )
                if es_estudio:
                    result["estudioBiblico"] = {
                        "titulo":   titulo,
                        "duracion": dur or 30,
                        "conductor": match_pub(fmt_nombre(val_b)),
                        "lector":    None,
                    }
                else:
                    result["vidaCristiana"].append({
                        "titulo":   titulo,
                        "tipo":     "parte",
                        "duracion": dur,
                        "pubId":    match_pub(fmt_nombre(val_b)),
                    })
                    vida_count += 1
            continue

    # Fallback: si no hay ministerio, crear lista vacía con una parte
    if not result["ministerio"]:
        result["ministerio"] = [{"titulo": "", "tipo": "demostracion", "duracion": None, "pubId": None, "ayudante": None}]
    if not result["vidaCristiana"]:
        result["vidaCristiana"] = [{"titulo": "", "tipo": "parte", "duracion": None, "pubId": None}]

    return result


# ─── Parser de hoja completa ──────────────────────────────────────────────────

def parse_hoja(ws, nombre_hoja):
    """
    Devuelve lista de dicts {fecha, ...} con todos los programas semanales de la hoja.
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # Fila 2 (idx 1): "Sala Auxiliar: Nombre"
    conductor_aux_mes = None
    if len(rows) > 1:
        celda_a1 = celda(rows[1], 0)
        if celda_a1:
            m_aux = re.search(r"Sala\s+Auxiliar[:\s]+(.+)", str(celda_a1), re.IGNORECASE)
            if m_aux:
                conductor_aux_mes = fmt_nombre(m_aux.group(1).strip())

    # Detectar bloques semanales
    semanas = []
    bloque_inicio = None
    bloque_fecha  = None

    for i, row in enumerate(rows):
        label = str(celda(row, 0) or "").strip()
        if label.lower().startswith("semana del"):
            fecha = parse_semana_str(label)
            if fecha:
                if bloque_inicio is not None and bloque_fecha is not None:
                    # Cerrar bloque anterior
                    bloque_rows = rows[bloque_inicio + 1 : i]
                    programa = parse_block(bloque_rows)
                    programa["fecha"]            = bloque_fecha
                    programa["conductorAuxMes"]  = conductor_aux_mes
                    semanas.append(programa)
                bloque_inicio = i
                bloque_fecha  = fecha

    # Último bloque
    if bloque_inicio is not None and bloque_fecha is not None:
        bloque_rows = rows[bloque_inicio + 1:]
        # Filtrar filas vacías del final
        while bloque_rows and all(v is None for v in bloque_rows[-1]):
            bloque_rows.pop()
        programa = parse_block(bloque_rows)
        programa["fecha"]           = bloque_fecha
        programa["conductorAuxMes"] = conductor_aux_mes
        semanas.append(programa)

    return semanas


# ─── Main ─────────────────────────────────────────────────────────────────────

print("==> Conectando a Firebase...")
cred = credentials.Certificate("serviceAccountKey.json")
firebase_admin.initialize_app(cred)
db = firestore.client()
congre_ref = db.collection("congregaciones").document(CONGRE_ID)

print(f"==> Cargando publicadores de '{CONGRE_ID}'...")
pubs_snap = congre_ref.collection("publicadores").stream()
for pub in pubs_snap:
    data = pub.to_dict()
    nombre = data.get("nombre", "")
    if nombre:
        nombre_to_id[normalize(nombre)] = pub.id
print(f"    {len(nombre_to_id)} publicadores cargados.")

print(f"\n==> Abriendo Excel: {EXCEL_FILE}")
wb = load_workbook(EXCEL_FILE, read_only=True, data_only=True)

total_semanas   = 0
total_subidas   = 0
hojas_procesadas = []

for nombre_hoja in wb.sheetnames:
    # Ignorar hojas SR y Anc y SM
    if nombre_hoja in HOJAS_IGNORAR or nombre_hoja.strip().endswith(SUFIJO_IGNORAR):
        print(f"  [skip] '{nombre_hoja}'")
        continue

    ws = wb[nombre_hoja]
    semanas = parse_hoja(ws, nombre_hoja)
    print(f"\n  Hoja '{nombre_hoja}': {len(semanas)} semanas")

    # Subir en lotes de 400
    batch   = db.batch()
    batch_n = 0

    for s in semanas:
        fecha = s.get("fecha")
        if not fecha:
            continue
        ref = congre_ref.collection("vidaministerio").document(fecha)
        batch.set(ref, s)
        batch_n += 1
        total_semanas += 1
        if batch_n == 400:
            batch.commit()
            total_subidas += batch_n
            batch   = db.batch()
            batch_n = 0
        print(f"    {fecha}", end="")
        # Breve resumen de la semana
        n_min = len(s.get("ministerio", []))
        n_vc  = len(s.get("vidaCristiana", []))
        print(f"  min={n_min} vc={n_vc}")

    if batch_n:
        batch.commit()
        total_subidas += batch_n

    hojas_procesadas.append(nombre_hoja)

# ─── Resumen ──────────────────────────────────────────────────────────────────
print(f"\n{'='*55}")
print(f"IMPORTACIÓN COMPLETA")
print(f"  Hojas procesadas : {len(hojas_procesadas)}")
print(f"  Semanas subidas  : {total_subidas}")
print(f"{'='*55}")

if unmatched_log:
    print(f"\nNombres NO encontrados en publicadores ({len(unmatched_log)}):")
    for n in sorted(unmatched_log):
        print(f"  · {n}")
else:
    print("\nTodos los nombres fueron encontrados en publicadores.")
