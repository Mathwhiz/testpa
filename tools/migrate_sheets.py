"""
migrate_sheets.py
Migra data de TerritoryApp_JW.xlsx y Asignaciones_Sur.xlsx a Firebase Firestore.

Uso:
    pip install firebase-admin openpyxl
    python migrate_sheets.py

Archivos necesarios en la misma carpeta:
    - migrate_sheets.py
    - serviceAccountKey.json    (descargado de Firebase Console)
    - TerritoryApp_JW.xlsx
    - Asignaciones_Sur.xlsx     (renombrado sin acentos)
"""

import firebase_admin
from firebase_admin import credentials, firestore
from openpyxl import load_workbook
from datetime import datetime
import re

# ─── CONFIG ───────────────────────────────────────────────────────────────────

CONGRE_ID     = "sur"
CONGRE_NOMBRE = "Congregación Sur"
PIN_ENCARGADO = "1234"

GRUPOS = [
    {"id": "1", "color": "#378ADD", "pin": "1111"},
    {"id": "2", "color": "#EF9F27", "pin": "2222"},
    {"id": "3", "color": "#97C459", "pin": "3333"},
    {"id": "4", "color": "#D85A30", "pin": "4444"},
    {"id": "C", "color": "#7F77DD", "pin": "5555",
     "esCongreacion": True, "zoomId": "844 0225 6636", "clave": "479104"},
]

TERRITORIOS_ESPECIALES = {
    131: "no_predica",
    11:  "peligroso",
}

HOJA_GRUPO = {
    "Grupo 1":      "1",
    "Grupo 2":      "2",
    "Grupo 3":      "3",
    "Grupo 4":      "4",
    "Congregación": "C",
}

ROLES_MAP = {
    "LECTOR":               "LECTOR",
    "SONIDO 1":             "SONIDO_1",
    "SONIDO 2":             "SONIDO_2",
    "PLATAFORMA":           "PLATAFORMA",
    "MICROFONISTAS 1":      "MICROFONISTAS_1",
    "MICROFONISTAS 2":      "MICROFONISTAS_2",
    "ACOMODADOR AUDITORIO": "ACOMODADOR_AUDITORIO",
    "ACOMODADOR ENTRADA":   "ACOMODADOR_ENTRADA",
    "PRESIDENTE":           "PRESIDENTE",
    "REVISTAS":             "REVISTAS",
    "PUBLICACIONES":        "PUBLICACIONES",
}

# ─── HELPERS ──────────────────────────────────────────────────────────────────

def fmt_fecha(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if s in ("—", "-", "", "None"):
        return None
    m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{2,4})$", s)
    if m:
        d, mo, y = m.groups()
        if len(y) == 2:
            y = "20" + y
        return f"{y}-{mo.zfill(2)}-{d.zfill(2)}"
    return None

def fmt_nombre(val):
    if val is None:
        return None
    s = str(val).strip()
    if s in ("—", "-", ""):
        return None
    return s.title()

def terr_num(val):
    if val is None:
        return None
    m = re.search(r"(\d+)", str(val))
    return int(m.group(1)) if m else None

def safe_get(row, i):
    return row[i] if row and i < len(row) else None

# ─── INIT FIREBASE ────────────────────────────────────────────────────────────

print("🔥 Conectando a Firebase...")
cred = credentials.Certificate("serviceAccountKey.json")
firebase_admin.initialize_app(cred)
db = firestore.client()
congre_ref = db.collection("congregaciones").document(CONGRE_ID)

# ─── 1. CONFIG ────────────────────────────────────────────────────────────────

print(f"\n📋 Creando config de congregación '{CONGRE_ID}'...")
congre_ref.set({
    "nombre":       CONGRE_NOMBRE,
    "pinEncargado": PIN_ENCARGADO,
    "creadoEn":     firestore.SERVER_TIMESTAMP,
})
for g in GRUPOS:
    congre_ref.collection("grupos").document(g["id"]).set(g)
    print(f"   ✓ Grupo {g['id']}")

# ─── 2. PUBLICADORES ──────────────────────────────────────────────────────────

print("\n👥 Importando publicadores...")
wb_asig    = load_workbook("Asignaciones_Sur.xlsx", read_only=True)
ws_lista   = wb_asig["Lista"]
rows_lista = list(ws_lista.iter_rows(values_only=True))

pub_count = 0
for row in rows_lista[2:]:
    if not row or not safe_get(row, 0):
        continue
    nombre    = str(row[0]).strip()
    roles_raw = str(safe_get(row, 1) or "").strip()
    roles     = [r.strip() for r in roles_raw.split(",") if r.strip()]
    congre_ref.collection("publicadores").document().set({
        "nombre": nombre,
        "roles":  roles,
        "activo": True,
    })
    pub_count += 1

print(f"   ✓ {pub_count} publicadores importados")

# ─── 3. ASIGNACIONES ──────────────────────────────────────────────────────────

print("\n📅 Importando asignaciones de reunión...")
ws_prog   = wb_asig["Programacion"]
rows_prog = list(ws_prog.iter_rows(values_only=True))
headers   = rows_prog[1] if len(rows_prog) > 1 else []

asig_count = 0
for row in rows_prog[2:]:
    if not row or not safe_get(row, 0):
        continue
    fecha = fmt_fecha(safe_get(row, 0))
    if not fecha:
        continue
    dia   = str(safe_get(row, 1) or "").strip()
    roles = {}
    for i, h in enumerate(headers):
        if h and str(h).strip() in ROLES_MAP:
            val = safe_get(row, i)
            if val:
                roles[ROLES_MAP[str(h).strip()]] = str(val).strip()
    congre_ref.collection("asignaciones").document().set({
        "fecha":     fecha,
        "diaSemana": dia,
        "roles":     roles,
    })
    asig_count += 1

print(f"   ✓ {asig_count} asignaciones importadas")

# ─── 4. TERRITORIOS E HISTORIAL ───────────────────────────────────────────────

print("\n🗺️  Importando territorios e historial...")
wb_terr    = load_workbook("TerritoryApp_JW.xlsx", read_only=True)
total_terr = 0
total_hist = 0

for hoja, grupo_id in HOJA_GRUPO.items():
    ws   = wb_terr[hoja]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        continue

    terr_row  = rows[0]
    terr_cols = []
    for i, val in enumerate(terr_row):
        if val and str(val).startswith("Terr."):
            num = terr_num(val)
            if num:
                terr_cols.append((i, num))

    for col_idx, terr_id in terr_cols:
        tipo     = TERRITORIOS_ESPECIALES.get(terr_id, "normal")
        terr_ref = congre_ref.collection("territorios").document(str(terr_id))
        terr_ref.set({
            "id":      terr_id,
            "grupoId": grupo_id,
            "tipo":    tipo,
        })
        total_terr += 1

        for data_row in rows[2:]:
            conductor_raw = safe_get(data_row, col_idx)
            inicio_raw    = safe_get(data_row, col_idx + 1)
            fin_raw       = safe_get(data_row, col_idx + 2)

            if conductor_raw is None and inicio_raw is None:
                continue

            conductor = fmt_nombre(conductor_raw)
            inicio    = fmt_fecha(inicio_raw)
            fin       = fmt_fecha(fin_raw)

            if not inicio:
                continue

            terr_ref.collection("historial").document().set({
                "conductor":   conductor,
                "fechaInicio": inicio,
                "fechaFin":    fin,
            })
            total_hist += 1

    print(f"   ✓ Grupo {grupo_id}: {len(terr_cols)} territorios")

print(f"   Total territorios : {total_terr}")
print(f"   Total historial   : {total_hist} entradas")

# ─── 5. SALIDAS ───────────────────────────────────────────────────────────────

print("\n📤 Importando salidas registradas...")
ws_hist3  = wb_terr["Historial 3"]
rows_hist = [r for r in ws_hist3.iter_rows(values_only=True) if any(v is not None for v in r)]

salida_count = 0
for row in rows_hist:
    fecha_reg = fmt_fecha(safe_get(row, 0))
    if not fecha_reg:
        continue
    notas     = str(safe_get(row, 1) or "").strip() or None
    terr_raw  = safe_get(row, 2)
    fecha_sal = fmt_fecha(safe_get(row, 3))
    conductor = fmt_nombre(safe_get(row, 4))
    hora      = str(safe_get(row, 5) or "").strip() or None

    if isinstance(terr_raw, float):
        terr_val = int(terr_raw)
    else:
        terr_val = str(terr_raw).strip() if terr_raw else None

    congre_ref.collection("salidas").document().set({
        "fechaRegistro": fecha_reg,
        "fechaSalida":   fecha_sal,
        "territorioId":  terr_val,
        "conductor":     conductor,
        "hora":          hora,
        "notas":         notas if notas and notas != "—" else None,
    })
    salida_count += 1

print(f"   ✓ {salida_count} salidas importadas")

# ─── RESUMEN ──────────────────────────────────────────────────────────────────

print("\n" + "=" * 50)
print("✅ Migración completa!")
print(f"   Congregación : {CONGRE_NOMBRE}")
print(f"   Grupos       : {len(GRUPOS)}")
print(f"   Publicadores : {pub_count}")
print(f"   Territorios  : {total_terr}")
print(f"   Historial    : {total_hist} entradas")
print(f"   Asignaciones : {asig_count}")
print(f"   Salidas      : {salida_count}")
print("=" * 50)